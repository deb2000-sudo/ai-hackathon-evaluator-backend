"""
Export hackathon submission data to Excel for admins.

One worksheet per timeline round. All submissions are included (with or without
video). Team rounds include full roster; solo rounds show the submitter only.
"""

from __future__ import annotations

import io
import json
import re
from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.exceptions import NotFoundError
from app.services.firebase import FirebaseService
from app.services.hackathon_service import HackathonService
from app.services.submission.base import SubmissionServiceBase
from app.services.team_service import TEAMS
from app.services.user_service import UserService
from app.utils.hackathon_round import normalize_max_team_size, round_title
from app.utils.time import now_ist_iso


EXPORT_COLUMNS: tuple[str, ...] = (
    "Submission ID",
    "Round Index",
    "Round Title",
    "Participation Mode",
    "Team Name",
    "Hackathon Team ID",
    "Submitter Student ID",
    "Submitter Name",
    "Submitter Email",
    "Submitter NIAT ID",
    "Submitter Mobile",
    "Submitter University",
    "Member 1 Name",
    "Member 1 Email",
    "Member 1 Role",
    "Member 2 Name",
    "Member 2 Email",
    "Member 2 Role",
    "Member 3 Name",
    "Member 3 Email",
    "Member 3 Role",
    "Member 4 Name",
    "Member 4 Email",
    "Member 4 Role",
    "Theme ID",
    "Theme Name",
    "Problem Statement",
    "Solution Description",
    "MVP Link",
    "GitHub Link",
    "Additional Field Answers (JSON)",
    "Demo Video Required",
    "Has Video",
    "Video GCS Path",
    "Video Source",
    "Video Filename",
    "Submission Status",
    "Review Status",
    "Final Score",
    "Report Published",
    "Assigned Evaluator",
    "Evaluator Notes",
    "Review Notes",
    "Auto AI Evaluation",
    "Submitted At",
    "Updated At",
)

_SHEET_BAD_CHARS = re.compile(r"[\[\]:*?/\\]")


@dataclass(frozen=True)
class ExportSheetData:
    title: str
    headers: list[str]
    rows: list[list[Any]]


@dataclass(frozen=True)
class HackathonExportDataset:
    hackathon: dict[str, Any]
    submission_count: int
    sheets: list[ExportSheetData]


class HackathonExportService:
    def __init__(
        self,
        firebase: FirebaseService | None = None,
        hackathon_service: HackathonService | None = None,
        user_service: UserService | None = None,
    ):
        self.firebase = firebase or FirebaseService()
        self.hackathon_service = hackathon_service or HackathonService(
            firebase=self.firebase
        )
        self.user_service = user_service or UserService(firebase=self.firebase)
        self.submissions_collection = SubmissionServiceBase.collection

    def build_export_workbook(self, hackathon_id: str) -> tuple[bytes, str]:
        """
        Build an ``.xlsx`` workbook for one hackathon.

        Returns ``(file_bytes, suggested_filename)``.
        """
        dataset = self.build_export_dataset(hackathon_id)
        hackathon = dataset.hackathon

        workbook = Workbook()
        workbook.remove(workbook.active)
        for sheet_data in dataset.sheets:
            sheet = workbook.create_sheet(title=sheet_data.title)
            if sheet_data.title == "Summary" and not sheet_data.headers:
                for row in sheet_data.rows:
                    sheet.append(row)
                sheet["A1"].font = Font(bold=True)
            else:
                self._write_sheet_values(sheet, sheet_data.headers, sheet_data.rows)

        buffer = io.BytesIO()
        workbook.save(buffer)
        safe_name = self._safe_filename_part(hackathon.get("name") or "hackathon")
        filename = f"{safe_name}_submissions_{now_ist_iso()[:10]}.xlsx"
        return buffer.getvalue(), filename

    def build_export_dataset(self, hackathon_id: str) -> HackathonExportDataset:
        hackathon = self.hackathon_service.get_hackathon(hackathon_id.strip())
        if not hackathon:
            raise NotFoundError("Hackathon not found", code="HACKATHON_NOT_FOUND")

        timeline = hackathon.get("timeline") or []
        submissions = self._submissions_for_hackathon(hackathon_id)
        by_round: dict[int, list[dict[str, Any]]] = {}
        for submission in submissions:
            round_index = int(submission.get("round_index") or 0)
            by_round.setdefault(round_index, []).append(submission)

        sheets: list[ExportSheetData] = [
            ExportSheetData(
                title="Summary",
                headers=[],
                rows=self._summary_rows(hackathon, len(submissions)),
            )
        ]

        if not timeline:
            rows = [
                self._build_row(submission, hackathon, 0)
                for submission in sorted(
                    submissions, key=lambda item: item.get("created_at") or ""
                )
            ]
            sheets.append(
                ExportSheetData(
                    title="Submissions",
                    headers=list(EXPORT_COLUMNS),
                    rows=rows,
                )
            )
        else:
            used_titles: set[str] = {"Summary"}
            for index, _round in enumerate(timeline):
                title = round_title(hackathon, index)
                sheet_name = self._unique_sheet_name(title, index, used_titles)
                round_submissions = sorted(
                    by_round.get(index, []),
                    key=lambda item: item.get("created_at") or "",
                )
                sheets.append(
                    ExportSheetData(
                        title=sheet_name,
                        headers=list(EXPORT_COLUMNS),
                        rows=[
                            self._build_row(submission, hackathon, index)
                            for submission in round_submissions
                        ],
                    )
                )

        return HackathonExportDataset(
            hackathon=hackathon,
            submission_count=len(submissions),
            sheets=sheets,
        )

    def _submissions_for_hackathon(self, hackathon_id: str) -> list[dict[str, Any]]:
        return self.firebase.query_collection(
            self.submissions_collection,
            "hackathon_id",
            "==",
            hackathon_id.strip(),
        )

    def _summary_rows(
        self, hackathon: dict[str, Any], export_count: int
    ) -> list[list[Any]]:
        return [
            ["Hackathon", hackathon.get("name") or ""],
            ["Hackathon ID", hackathon.get("id") or ""],
            ["Start Date", hackathon.get("start_date") or ""],
            ["End Date", hackathon.get("end_date") or ""],
            ["Exported At (IST)", now_ist_iso()],
            ["Total Submissions Exported", export_count],
            [
                "Note",
                "All submissions are included. Video GCS path is blank when no demo video.",
            ],
        ]

    def _write_sheet_values(
        self,
        sheet,
        headers: list[str],
        rows: list[list[Any]],
    ) -> None:
        if headers:
            sheet.append(headers)
            for cell in sheet[1]:
                cell.font = Font(bold=True)
        for row in rows:
            sheet.append(row)
        width_columns = headers or (rows[0] if rows else [])
        for index, _ in enumerate(width_columns, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = 18

    def _build_row(
        self,
        submission: dict[str, Any],
        hackathon: dict[str, Any],
        round_index: int,
    ) -> list[Any]:
        _, round_title_text, max_team_size = self._round_meta(hackathon, round_index)
        team_doc, team_id = self._load_team(submission.get("hackathon_team_id"))
        participation_mode = "Solo" if max_team_size <= 1 else "Team"

        submitter = self._profile(submission.get("student_id"))
        team_name = str(submission.get("team_name") or "").strip()
        if team_doc:
            team_name = str(team_doc.get("team_name") or team_name or "Team")
        elif not team_name and submitter:
            team_name = str(submitter.get("name") or "")

        members = self._team_members(team_doc, submitter, participation_mode)
        member_cells: list[str] = []
        for slot in range(4):
            if slot < len(members):
                member_cells.extend(
                    [
                        members[slot].get("name") or "",
                        members[slot].get("email") or "",
                        members[slot].get("role") or "",
                    ]
                )
            else:
                member_cells.extend(["", "", ""])

        field_answers = submission.get("field_answers") or {}
        if not isinstance(field_answers, dict):
            field_answers = {}

        problem = self._submission_text(submission, field_answers, "problem_statement")
        solution = self._submission_text(
            submission, field_answers, "solution_description"
        )
        mvp_link = self._submission_text(submission, field_answers, "mvp_link")
        github_link = self._submission_text(
            submission,
            field_answers,
            "github_link",
            fallback_keys=("project_github_link",),
        )
        extra_answers = self._extra_field_answers_json(
            field_answers,
            reserved={
                "problem_statement",
                "solution_description",
                "mvp_link",
                "github_link",
                "project_github_link",
            },
        )

        video_path = str(submission.get("video_path") or "").strip()
        has_video = bool(video_path.startswith("gs://"))

        return [
            submission.get("id") or "",
            submission.get("round_index", round_index),
            submission.get("round_title") or round_title_text,
            participation_mode,
            team_name,
            team_id or submission.get("hackathon_team_id") or "",
            submission.get("student_id") or "",
            submitter.get("name") if submitter else "",
            submitter.get("email") if submitter else "",
            submitter.get("niat_id") if submitter else "",
            submitter.get("mobile_no") if submitter else "",
            submitter.get("university") if submitter else "",
            *member_cells,
            submission.get("theme_id") or "",
            submission.get("theme_name") or "",
            problem,
            solution,
            mvp_link,
            github_link,
            extra_answers,
            self._yes_no(submission.get("working_demo_video_required", True)),
            self._yes_no(has_video),
            video_path if has_video else "",
            submission.get("video_source") or "",
            submission.get("source_filename") or "",
            submission.get("status") or "",
            submission.get("review_status") or "",
            submission.get("final_score"),
            self._yes_no(submission.get("report_published")),
            submission.get("assigned_evaluator_name") or "",
            submission.get("evaluator_notes") or "",
            submission.get("review_notes") or "",
            self._yes_no(submission.get("auto_ai_evaluation")),
            submission.get("created_at") or "",
            submission.get("updated_at") or "",
        ]

    @staticmethod
    def _submission_text(
        submission: dict[str, Any],
        field_answers: dict[str, Any],
        key: str,
        *,
        fallback_keys: tuple[str, ...] = (),
    ) -> str:
        direct = submission.get(key)
        if direct is not None and str(direct).strip():
            return str(direct).strip()
        from_answers = field_answers.get(key)
        if from_answers is not None and str(from_answers).strip():
            return str(from_answers).strip()
        for alt in fallback_keys:
            alt_val = field_answers.get(alt)
            if alt_val is not None and str(alt_val).strip():
                return str(alt_val).strip()
        return ""

    @staticmethod
    def _extra_field_answers_json(
        field_answers: dict[str, Any],
        *,
        reserved: set[str],
    ) -> str:
        extra = {
            key: value
            for key, value in field_answers.items()
            if key not in reserved and value not in (None, "")
        }
        if not extra:
            return ""
        return json.dumps(extra, ensure_ascii=False)

    @staticmethod
    def _yes_no(value: Any) -> str:
        return "Yes" if bool(value) else "No"

    def _round_meta(
        self, hackathon: dict[str, Any], round_index: int
    ) -> tuple[dict[str, Any], str, int]:
        timeline = hackathon.get("timeline") or []
        if round_index < 0 or round_index >= len(timeline):
            return {}, f"Round {round_index + 1}", 1
        round_ = timeline[round_index]
        if not isinstance(round_, dict):
            round_ = dict(round_)
        title = str(round_.get("title") or f"Round {round_index + 1}")
        max_size = normalize_max_team_size(round_.get("max_team_size", 1))
        return round_, title, max_size

    def _load_team(
        self, team_id: str | None
    ) -> tuple[dict[str, Any] | None, str | None]:
        if not team_id:
            return None, None
        doc = self.firebase.get_document(TEAMS, str(team_id))
        return (doc, str(team_id)) if doc else (None, str(team_id))

    def _profile(self, user_id: str | None) -> dict[str, Any]:
        if not user_id:
            return {}
        return self.user_service.get_user(str(user_id)) or {}

    def _team_members(
        self,
        team_doc: dict[str, Any] | None,
        submitter: dict[str, Any],
        participation_mode: str,
    ) -> list[dict[str, str]]:
        if participation_mode == "Solo":
            if not submitter:
                return []
            return [
                {
                    "name": str(submitter.get("name") or ""),
                    "email": str(submitter.get("email") or ""),
                    "role": "solo",
                }
            ]

        if team_doc:
            members = team_doc.get("members") or []
            return [
                {
                    "name": str(member.get("name") or ""),
                    "email": str(member.get("email") or ""),
                    "role": str(member.get("role") or "member"),
                }
                for member in members
            ]

        if submitter:
            return [
                {
                    "name": str(submitter.get("name") or ""),
                    "email": str(submitter.get("email") or ""),
                    "role": "leader",
                }
            ]
        return []

    @staticmethod
    def _unique_sheet_name(title: str, index: int, used: set[str]) -> str:
        base = _SHEET_BAD_CHARS.sub("", title).strip() or f"Round {index + 1}"
        base = base[:31]
        candidate = base
        suffix = 2
        while candidate in used:
            tail = f" {suffix}"
            candidate = f"{base[: 31 - len(tail)]}{tail}"
            suffix += 1
        used.add(candidate)
        return candidate

    @staticmethod
    def _safe_filename_part(value: str) -> str:
        cleaned = re.sub(r"[^\w\s-]", "", value).strip().replace(" ", "_")
        return cleaned[:60] or "hackathon"
