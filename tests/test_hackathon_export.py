"""Admin hackathon submission Excel export."""

import io

import pytest
from openpyxl import load_workbook

from app.exceptions import NotFoundError
from app.services.hackathon_export_service import (
    EXPORT_COLUMNS,
    HackathonExportService,
)


class FakeFirebase:
    def __init__(self, submissions: list[dict], teams: dict[str, dict] | None = None) -> None:
        self.submissions = submissions
        self.teams = teams or {}

    def query_collection(self, collection, field, operator, value):
        if collection != "submissions" or field != "hackathon_id":
            return []
        return [
            item for item in self.submissions if item.get("hackathon_id") == value
        ]

    def get_document(self, collection, document_id):
        if collection == "hackathon_teams":
            doc = self.teams.get(document_id)
            return dict(doc) if doc else None
        return None


class FakeHackathonService:
    def __init__(self, hackathon: dict | None) -> None:
        self.hackathon = hackathon

    def get_hackathon(self, hackathon_id: str):
        if not self.hackathon or self.hackathon.get("id") != hackathon_id:
            return None
        return dict(self.hackathon)


class FakeUserService:
    def __init__(self, users: dict[str, dict]) -> None:
        self.users = users

    def get_user(self, user_id: str):
        return self.users.get(user_id)


def _service(
    *,
    hackathon: dict | None,
    submissions: list[dict],
    users: dict[str, dict] | None = None,
    teams: dict[str, dict] | None = None,
) -> HackathonExportService:
    return HackathonExportService(
        firebase=FakeFirebase(submissions, teams),
        hackathon_service=FakeHackathonService(hackathon),
        user_service=FakeUserService(users or {}),
    )


def test_export_includes_all_submissions_and_full_fields():
    hackathon = {
        "id": "hack-1",
        "name": "Summer Hack",
        "start_date": "2026-09-01",
        "end_date": "2026-09-30",
        "timeline": [
            {"title": "Round 1", "max_team_size": 2},
            {"title": "Round 2", "max_team_size": 1},
        ],
    }
    submissions = [
        {
            "id": "sub-team",
            "hackathon_id": "hack-1",
            "round_index": 0,
            "round_title": "Round 1",
            "student_id": "leader-1",
            "hackathon_team_id": "team-1",
            "team_name": "Alpha",
            "theme_id": "theme-1",
            "theme_name": "AI",
            "problem_statement": "Problem A",
            "solution_description": "Solution A",
            "mvp_link": "https://demo.example.com",
            "github_link": "https://github.com/example/a",
            "field_answers": {"custom_metric": "High"},
            "video_path": "gs://bucket/videos/team.mp4",
            "video_source": "uploaded",
            "source_filename": "demo.mp4",
            "status": "uploaded",
            "review_status": "none",
            "working_demo_video_required": True,
            "created_at": "2026-09-10T10:00:00+05:30",
            "updated_at": "2026-09-10T10:00:00+05:30",
        },
        {
            "id": "sub-no-video",
            "hackathon_id": "hack-1",
            "round_index": 0,
            "round_title": "Round 1",
            "student_id": "solo-2",
            "theme_name": "Web",
            "problem_statement": "Text only problem",
            "solution_description": "Text only solution",
            "mvp_link": "https://mvp.example.com",
            "working_demo_video_required": False,
            "video_path": None,
            "status": "uploaded",
            "created_at": "2026-09-11T10:00:00+05:30",
            "updated_at": "2026-09-11T10:00:00+05:30",
        },
        {
            "id": "sub-solo",
            "hackathon_id": "hack-1",
            "round_index": 1,
            "round_title": "Round 2",
            "student_id": "solo-1",
            "team_name": "Solo Student",
            "theme_name": "Mobile",
            "problem_statement": "P2",
            "solution_description": "S2",
            "github_link": "https://github.com/example/solo",
            "video_path": "gs://bucket/videos/solo.mp4",
            "created_at": "2026-09-12T10:00:00+05:30",
            "updated_at": "2026-09-12T10:00:00+05:30",
        },
    ]
    users = {
        "leader-1": {
            "name": "Leader One",
            "email": "leader@example.com",
            "niat_id": "N1",
            "mobile_no": "+919999999999",
            "university": "NIAT",
        },
        "solo-1": {
            "name": "Solo Student",
            "email": "solo@example.com",
            "niat_id": "N2",
            "mobile_no": "+918888888888",
            "university": "NIAT",
        },
        "solo-2": {
            "name": "Text Only",
            "email": "text@example.com",
            "niat_id": "N3",
        },
    }
    teams = {
        "team-1": {
            "team_name": "Alpha Squad",
            "members": [
                {
                    "user_id": "leader-1",
                    "name": "Leader One",
                    "email": "leader@example.com",
                    "role": "leader",
                },
                {
                    "user_id": "member-1",
                    "name": "Member One",
                    "email": "member@example.com",
                    "role": "member",
                },
            ],
        }
    }
    svc = _service(
        hackathon=hackathon,
        submissions=submissions,
        users=users,
        teams=teams,
    )
    content, filename = svc.build_export_workbook("hack-1")
    assert filename.endswith(".xlsx")

    workbook = load_workbook(io.BytesIO(content))
    round1 = workbook["Round 1"]
    headers = [cell.value for cell in round1[1]]
    assert headers == list(EXPORT_COLUMNS)

    round1_rows = list(round1.iter_rows(min_row=2, values_only=True))
    assert len(round1_rows) == 2

    team_row = next(row for row in round1_rows if row[0] == "sub-team")
    assert team_row[headers.index("Problem Statement")] == "Problem A"
    assert team_row[headers.index("MVP Link")] == "https://demo.example.com"
    assert team_row[headers.index("GitHub Link")] == "https://github.com/example/a"
    assert team_row[headers.index("Has Video")] == "Yes"
    assert team_row[headers.index("Video GCS Path")] == "gs://bucket/videos/team.mp4"
    assert '"custom_metric"' in team_row[headers.index("Additional Field Answers (JSON)")]

    text_row = next(row for row in round1_rows if row[0] == "sub-no-video")
    assert text_row[headers.index("Has Video")] == "No"
    assert (text_row[headers.index("Video GCS Path")] or "") == ""
    assert text_row[headers.index("MVP Link")] == "https://mvp.example.com"


def test_export_missing_hackathon_raises():
    svc = _service(hackathon=None, submissions=[])
    with pytest.raises(NotFoundError) as exc:
        svc.build_export_workbook("missing")
    assert exc.value.code == "HACKATHON_NOT_FOUND"
